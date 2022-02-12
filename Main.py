
#1. accesing worksheets and cell values-------------------------------

from json import load
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
""" wb = load_workbook("Grades.xlsx")
ws = wb.active
print(ws['A2'].value)

#change the value of a cell
ws['A2'].value = 'Test'

#save the workbook
wb.save("Grades.xlsx") """

#2. creating,listing and changing sheets---------------------------------

#print sheet names of thw workbook
""" print(wb.sheetnames)

print(wb['Sheet1'])

#create a sheet
wb = load_workbook("Grades.xlsx")
ws = wb.active
wb.create_sheet("Test")
print(wb.sheetnames) """


#3. creating a new workbook--------------------------------------

""" wb = Workbook()
ws = wb.active
#changing the title of the sheet
ws.title = "Data"

#4. appending or adding rows-------------------------------------
ws.append(['Dew', 'is', 'Great', '!'])
ws.append(['Dew', 'is', 'Great', '!'])
ws.append(['Dew', 'is', 'Great', '!'])
ws.append(['end'])
wb.save("Dew.xlsx") """

#5. accessing multiple rows-----------------------------------

""" wb = load_workbook("Dew.xlsx")
ws = wb.active
#loop through the data
for row in range(1,11):
    for col in range(1,5):
        char = get_column_letter(col)#taking numbers between 1 and 26
        ws[char + str(row)] = char + str(row)
wb.save("Dew.xlsx") """

#6. merging cells---------------------------

""" wb = load_workbook("Dew.xlsx")
ws = wb.active

ws.merge_cells("A1:D2")
wb.save("Dew.xlsx") """

#7. inserting and deleting rows
""" wb = load_workbook("Dew.xlsx")
ws = wb.active
ws.delete_cols(2)
wb.save("Dew.xlsx") """

#8. copying and moving cells
""" wb = load_workbook("Dew.xlsx")
ws = wb.active
ws.move_range("C1:D11", rows = 2, cols = 2)
wb.save("Dew.xlsx") """












