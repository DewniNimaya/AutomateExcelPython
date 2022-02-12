from pydoc import importfile
from openpyxl import Workbook, workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import fonts

#create xl using python code

data = {
    "Joe":{
        "math":65,
        "science":78,
        "english":98,
        "gym":89
    },
    "Bill":{
        "math":55,
        "science":72,
        "english":87,
        "gym":85
    },
    "Tim":{
        "math":100,
        "science":45,
        "english":75,
        "gym":97
    },
    "Sally":{
        "math":30,
        "science":25,
        "english":45,
        "gym":100
    },
    "Jane":{
        "math":100,
        "science":100,
        "english":100,
        "gym":60
    }
}

wb = Workbook()
ws=wb.active
ws.title = "Grades"

headings = ['Name'] + list(data['Joe'].keys())
ws.append(headings)

for person in data:
    grades = list(data[person].values())
    ws.append([person] + grades)

# calculating the average of all grades
for col in range(2, len(data['Joe'])+2):
    char = get_column_letter(col)
    ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

wb.save("NewGrades.xlsx")
